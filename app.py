#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
最终完整版：POS 订单支付 vs 订单菜品 对账（自动兼容 xls/xlsx/“内容是xls但后缀是xlsx”）

功能：
1) 弹出文件选择框：选择【订单支付报告】、【订单菜品报告】两个文件
2) 自动识别真实格式（看文件头，不看后缀）
   - OLE2/CFB(BIFF) => xls => xlrd 读取
   - ZIP => xlsx => openpyxl(read_only=True) 流式读取
3) 对账逻辑：
   - 正常订单：按 POS销售单号 对账（支付表“总金额” vs 菜品表“优惠后小计价格”汇总）
   - 退款订单：按 POS退款单号 对账（只要“POS退款单号”非空，则按退款单号聚合对账）
4) 容差：默认 1 角（0.10 元 = 10 分）
5) 输出：
   - 差异明细 CSV（中文表头）
   - 差异原因统计 CSV（中文表头）
   保存位置由你在“另存为”对话框自主选择

依赖：
pip install xlrd openpyxl pandas
"""

import os
import re
from collections import defaultdict, Counter
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd


# -----------------------------
# 文件选择/保存对话框
# -----------------------------
def pick_two_files() -> Tuple[str, str]:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    pay_path = filedialog.askopenfilename(
        title="请选择【订单支付报告】文件（xls/xlsx均可）",
        filetypes=[("Excel 文件", "*.xls *.xlsx"), ("所有文件", "*.*")]
    )
    if not pay_path:
        raise RuntimeError("未选择订单支付报告文件。")

    item_path = filedialog.askopenfilename(
        title="请选择【订单菜品报告】文件（xls/xlsx均可）",
        filetypes=[("Excel 文件", "*.xls *.xlsx"), ("所有文件", "*.*")]
    )
    if not item_path:
        raise RuntimeError("未选择订单菜品报告文件。")

    return pay_path, item_path


def pick_save_path(default_name: str, title: str) -> str:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    path = filedialog.asksaveasfilename(
        title=title,
        defaultextension=".csv",
        initialfile=default_name,
        filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")]
    )
    if not path:
        raise RuntimeError("未选择保存位置，已取消输出。")
    return path


# -----------------------------
# 格式识别（不看后缀，看文件头）
# -----------------------------
def detect_excel_format(path: str) -> str:
    """
    返回 'xls' 或 'xlsx'
    - xls (OLE2/CFB): D0 CF 11 E0 A1 B1 1A E1
    - xlsx (ZIP): PK 03 04 / PK 05 06 / PK 07 08
    """
    with open(path, "rb") as f:
        head = f.read(8)

    if head.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return "xls"
    if head.startswith(b"PK\x03\x04") or head.startswith(b"PK\x05\x06") or head.startswith(b"PK\x07\x08"):
        return "xlsx"

    raise RuntimeError(f"无法识别文件格式：{path}（既不是 OLE2 xls，也不是 ZIP xlsx）")


# -----------------------------
# 工具函数
# -----------------------------
def norm_str(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    # 单号常被读成浮点：去掉末尾 .0
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def money_to_cents(x: Any) -> int:
    """金额转“分”(int)，避免浮点误差。"""
    if x is None:
        return 0
    if isinstance(x, int):
        return x * 100
    if isinstance(x, float):
        return int(round(x * 100))
    s = str(x).strip().replace(",", "")
    if s == "":
        return 0
    try:
        return int(round(float(s) * 100))
    except ValueError:
        return 0


def cents_to_money_str(cents: int) -> str:
    sign = "-" if cents < 0 else ""
    cents = abs(cents)
    return f"{sign}{cents // 100}.{cents % 100:02d}"


def find_header_row_in_rows(rows: List[List[Any]], required_cols: List[str]) -> Optional[Tuple[int, Dict[str, int]]]:
    required = set(required_cols)
    for r, row in enumerate(rows):
        row_vals = [norm_str(v) for v in row]
        row_set = set([v for v in row_vals if v])
        if required.issubset(row_set):
            col_map = {name: row_vals.index(name) for name in required_cols}
            return r, col_map
    return None


def open_excel_iter(path: str, required_cols: List[str], scan_rows: int = 60):
    """
    统一打开并返回数据行迭代器（从表头下一行开始）：
    返回：header_row_index(0-based), col_map, rows_iterator
    """
    fmt = detect_excel_format(path)

    if fmt == "xls":
        import xlrd
        book = xlrd.open_workbook(path, formatting_info=False)
        sh = book.sheet_by_index(0)

        max_r = min(scan_rows, sh.nrows)
        sample = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(max_r)]
        found = find_header_row_in_rows(sample, required_cols)
        if not found:
            raise RuntimeError(f"未在前 {scan_rows} 行找到表头（xls）：{required_cols}")
        header_r, cmap = found

        def row_iter():
            for r in range(header_r + 1, sh.nrows):
                yield [sh.cell_value(r, c) for c in range(sh.ncols)]

        return header_r, cmap, row_iter()

    # fmt == "xlsx"
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    sample = []
    for row in ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True):
        sample.append(list(row))
    found = find_header_row_in_rows(sample, required_cols)
    if not found:
        wb.close()
        raise RuntimeError(f"未在前 {scan_rows} 行找到表头（xlsx）：{required_cols}")
    header_r, cmap = found  # 0-based

    def row_iter():
        # openpyxl 行号 1-based，所以数据从 header_r(0-based)+2 开始
        for row in ws.iter_rows(min_row=header_r + 2, values_only=True):
            yield list(row)

    # 不要提前 close，ws 还要迭代；脚本结束进程退出即可
    return header_r, cmap, row_iter()


# -----------------------------
# 聚合结构
# -----------------------------
@dataclass
class Agg:
    cents_sum: int = 0
    row_count: int = 0
    meta_counter: Counter = None

    def __post_init__(self):
        if self.meta_counter is None:
            self.meta_counter = Counter()


def agg_add(a: Agg, cents: int, meta: Optional[str] = None):
    a.cents_sum += cents
    a.row_count += 1
    if meta:
        a.meta_counter[meta] += 1


# -----------------------------
# 读取并聚合：支付表/菜品表
# -----------------------------
def read_payment_agg(path: str) -> Tuple[Dict[str, Agg], Dict[str, Agg]]:
    """
    支付表：
    - POS退款单号为空：按 POS销售单号 聚合
    - POS退款单号非空：按 POS退款单号 聚合
    金额字段：总金额
    """
    required = ["POS销售单号", "POS退款单号", "总金额", "支付类型", "门店"]
    _, cmap, rows = open_excel_iter(path, required_cols=required)

    sales_agg: Dict[str, Agg] = defaultdict(Agg)
    refund_agg: Dict[str, Agg] = defaultdict(Agg)

    for row in rows:
        def v(col):
            idx = cmap[col]
            return row[idx] if idx < len(row) else None

        pos_sale = norm_str(v("POS销售单号"))
        pos_refund = norm_str(v("POS退款单号"))
        cents = money_to_cents(v("总金额"))
        pay_type = norm_str(v("支付类型"))
        store = norm_str(v("门店"))

        if not pos_sale and not pos_refund:
            continue

        meta = f"{store}|{pay_type}" if (store or pay_type) else None

        if pos_refund:
            agg_add(refund_agg[pos_refund], cents, meta)
        else:
            agg_add(sales_agg[pos_sale], cents, meta)

    return sales_agg, refund_agg


def read_items_agg(path: str) -> Tuple[Dict[str, Agg], Dict[str, Agg]]:
    """
    菜品表：
    - POS退款单号为空：按 POS销售单号 聚合
    - POS退款单号非空：按 POS退款单号 聚合
    金额字段：优惠后小计价格（按单号求和）
    """
    required = ["POS销售单号", "POS退款单号", "优惠后小计价格", "单据类型", "菜品状态", "门店", "菜品名称"]
    _, cmap, rows = open_excel_iter(path, required_cols=required)

    sales_agg: Dict[str, Agg] = defaultdict(Agg)
    refund_agg: Dict[str, Agg] = defaultdict(Agg)

    for row in rows:
        def v(col):
            idx = cmap[col]
            return row[idx] if idx < len(row) else None

        pos_sale = norm_str(v("POS销售单号"))
        pos_refund = norm_str(v("POS退款单号"))
        cents = money_to_cents(v("优惠后小计价格"))

        doc_type = norm_str(v("单据类型"))
        dish_status = norm_str(v("菜品状态"))
        store = norm_str(v("门店"))
        dish_name = norm_str(v("菜品名称"))

        if not pos_sale and not pos_refund:
            continue

        meta = f"{store}|{doc_type}|{dish_status}|{dish_name}" if (store or doc_type or dish_status or dish_name) else None

        if pos_refund:
            agg_add(refund_agg[pos_refund], cents, meta)
        else:
            agg_add(sales_agg[pos_sale], cents, meta)

    return sales_agg, refund_agg


# -----------------------------
# 对账：差异输出 & 原因猜测
# -----------------------------
def compare_aggs(pay: Dict[str, Agg], item: Dict[str, Agg], kind: str, tolerance_cents: int) -> pd.DataFrame:
    keys = set(pay.keys()) | set(item.keys())
    out = []

    for k in keys:
        pa = pay.get(k)
        ia = item.get(k)

        ps = pa.cents_sum if pa else 0
        is_ = ia.cents_sum if ia else 0
        diff = ps - is_

        if abs(diff) <= tolerance_cents:
            continue

        if pa is None:
            reason = "仅菜品存在：支付缺失/未导出/单号不一致"
        elif ia is None:
            reason = "仅支付存在：菜品缺失/未导出/单号不一致"
        else:
            if kind == "REFUND":
                if (ps > 0 and is_ < 0) or (ps < 0 and is_ > 0):
                    reason = "退款正负号方向不一致（一个表正一个表负）"
                else:
                    reason = "退款金额不一致：可能部分退款/多次退款/抹零/口径差异"
            else:
                if pa.row_count > 1:
                    reason = "支付多笔汇总仍不一致：可能混合支付/重复/拆单"
                elif ia.row_count > 20:
                    reason = "菜品行很多仍不一致：可能退菜/作废/状态口径差异"
                else:
                    reason = "金额不一致：可能抹零/服务费/包装费/非菜品费用/口径差异"

        pay_meta = "; ".join([f"{m}*{c}" for m, c in (pa.meta_counter.most_common(3) if pa else [])])
        item_meta = "; ".join([f"{m}*{c}" for m, c in (ia.meta_counter.most_common(3) if ia else [])])

        out.append({
            "kind": kind,
            "key": k,
            "pay_sum": cents_to_money_str(ps),
            "item_sum": cents_to_money_str(is_),
            "diff(pay-item)": cents_to_money_str(diff),
            "pay_rows": pa.row_count if pa else 0,
            "item_rows": ia.row_count if ia else 0,
            "reason_guess": reason,
            "pay_meta_top3": pay_meta,
            "item_meta_top3": item_meta,
        })

    df = pd.DataFrame(out)
    if not df.empty:
        df["_abs"] = df["diff(pay-item)"].apply(money_to_cents).abs()
        df = df.sort_values("_abs", ascending=False).drop(columns=["_abs"])
    return df


# -----------------------------
# 中文表头映射
# -----------------------------
COLUMN_CN_MAP = {
    "kind": "单据类型（销售/退款）",
    "key": "POS单号",
    "pay_sum": "支付表总金额",
    "item_sum": "菜品表优惠后金额",
    "diff(pay-item)": "金额差异（支付-菜品）",
    "pay_rows": "支付记录行数",
    "item_rows": "菜品记录行数",
    "reason_guess": "差异原因判断",
    "pay_meta_top3": "支付侧关键信息（门店|支付类型 Top3）",
    "item_meta_top3": "菜品侧关键信息（门店|单据类型|状态|菜品 Top3）",
}


def main():
    pay_path, item_path = pick_two_files()

    for p in (pay_path, item_path):
        if not os.path.exists(p):
            raise FileNotFoundError(p)
        fmt = detect_excel_format(p)
        print(f"已选择文件：{p}\n  真实格式识别为：{fmt}\n")

    # 容差：1角 = 0.10元 = 10分
    tolerance_cents = 10

    print("读取并聚合支付表…")
    pay_sales, pay_refund = read_payment_agg(pay_path)

    print("读取并聚合菜品表…")
    item_sales, item_refund = read_items_agg(item_path)

    print("对账：正常订单（按 POS销售单号）…")
    df_sale = compare_aggs(pay_sales, item_sales, kind="SALE", tolerance_cents=tolerance_cents)

    print("对账：退款订单（按 POS退款单号）…")
    df_refund = compare_aggs(pay_refund, item_refund, kind="REFUND", tolerance_cents=tolerance_cents)

    out_all = pd.concat([df_sale, df_refund], ignore_index=True)

    # 中文表头
    out_all = out_all.rename(columns=COLUMN_CN_MAP)

    # 让用户选择保存位置
    diff_path = pick_save_path("对账差异明细.csv", "请选择【差异明细表】保存位置")
    reason_path = pick_save_path("对账差异原因统计.csv", "请选择【差异原因统计】保存位置")

    out_all.to_csv(diff_path, index=False, encoding="utf-8-sig")

    # 差异原因统计
    if out_all.empty:
        pd.DataFrame(columns=["差异原因判断", "数量"]).to_csv(reason_path, index=False, encoding="utf-8-sig")
        print("✅ 未发现差异（在容忍误差范围内）")
    else:
        stat = (out_all["差异原因判断"]
                .value_counts()
                .reset_index()
                .rename(columns={"index": "差异原因判断", "差异原因判断": "数量"}))
        stat.to_csv(reason_path, index=False, encoding="utf-8-sig")
        print(f"✅ 发现差异条目：{len(out_all)}")

    print(f"✅ 已保存差异明细：{diff_path}")
    print(f"✅ 已保存原因统计：{reason_path}")
    print(f"容差：0.10 元（1角）")


if __name__ == "__main__":
    main()
